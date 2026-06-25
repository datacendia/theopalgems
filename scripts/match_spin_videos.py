"""
Match inventory products to their WhatsApp turntable clips.

Reads "INVENTORY OPAL GRAND (2).xlsx" (sheet INVENTARIO) and uses the
"Video File name" column (L) to resolve each product to an actual .mp4 in
Whatsapp1/<category>/. Normalizes labels (lowercase, strip non-alphanumerics)
so sheet labels like "Necklace 3" / "Ring7" / "Braclet1" match real files like
necklace3.mp4 / Ring7.mp4 / braclet1.mp4.

Read-only: prints a coverage report and writes scripts/_spin_match.json.
Nothing is renamed, copied, or processed here.

Usage: python scripts/match_spin_videos.py
"""
import json
import os
import re
import sys

BASE = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
XLSX = os.path.join(BASE, 'INVENTORY OPAL GRAND (2).xlsx')
WHATSAPP = os.path.join(BASE, 'Whatsapp1')
OUT_JSON = os.path.join(BASE, 'scripts', '_spin_match.json')

# Sheet category -> Whatsapp1 folder (note the folder typos: Braclets, Earings)
CAT_FOLDER = {
    'necklace': 'Necklaces',
    'ring': 'Rings',
    'bracelet': 'Braclets',
    'earring': 'Earings',
}


def norm(s):
    """Lowercase and strip everything but a-z0-9 for fuzzy filename matching."""
    return re.sub(r'[^a-z0-9]', '', str(s or '').lower())


def clean_id(rid):
    nid = str(rid)
    if nid.endswith('.0'):
        nid = nid[:-2]
    return nid


def list_clips(folder_name):
    """Return {normalized_stem: filename} for .mp4 files in a Whatsapp1 subfolder."""
    folder = os.path.join(WHATSAPP, folder_name)
    out = {}
    if not os.path.isdir(folder):
        return out
    for f in os.listdir(folder):
        if f.lower().endswith('.mp4'):
            stem = os.path.splitext(f)[0]
            out.setdefault(norm(stem), f)
    return out


def main():
    try:
        import openpyxl
    except ImportError:
        print('ERROR: openpyxl not installed. Run: pip install openpyxl')
        sys.exit(1)

    if not os.path.exists(XLSX):
        print(f'ERROR: spreadsheet not found: {XLSX}')
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb['INVENTARIO'] if 'INVENTARIO' in wb.sheetnames else wb.active

    # Pre-index clips per category folder
    clips = {cat: list_clips(folder) for cat, folder in CAT_FOLDER.items()}
    used_files = {cat: set() for cat in CAT_FOLDER}

    rows = []
    matched = 0
    no_label = 0
    unmatched = []

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None or str(r[0]).strip() == '':
            continue
        rid = clean_id(r[0])
        name = (r[1] or '') if len(r) > 1 else ''
        cat_raw = (r[2] or '') if len(r) > 2 else ''
        label = (r[11] or '') if len(r) > 11 else ''
        catkey = str(cat_raw).strip().lower()
        label_str = str(label).strip()

        rec = {
            'id': rid,
            'name': str(name).strip(),
            'category': catkey,
            'label': label_str,
            'file': None,
            'matched': False,
        }

        if not label_str or label_str.lower() == 'none':
            no_label += 1
            rows.append(rec)
            continue

        folder_clips = clips.get(catkey, {})
        hit = folder_clips.get(norm(label_str))
        if hit:
            rec['file'] = f'Whatsapp1/{CAT_FOLDER[catkey]}/{hit}'
            rec['matched'] = True
            matched += 1
            used_files[catkey].add(norm(label_str))
        else:
            unmatched.append(rec)
        rows.append(rec)

    total_products = len(rows)
    with_label = total_products - no_label

    print('=' * 64)
    print(f'Spreadsheet : {os.path.basename(XLSX)}  (sheet: {ws.title})')
    print(f'Products    : {total_products}')
    print(f'With a video label : {with_label}')
    print(f'MATCHED to a clip  : {matched}')
    print(f'No / "None" label  : {no_label}')
    print(f'Label but NO file  : {len(unmatched)}')
    print('=' * 64)

    if unmatched:
        print('\nLABELS WITH NO MATCHING FILE:')
        for u in unmatched:
            print(f'  [{u["id"]}] {u["name"]} ({u["category"]}) -> "{u["label"]}"')

    print('\nUNUSED CLIPS (present in folder, not referenced by any product):')
    for cat, folder in CAT_FOLDER.items():
        all_keys = set(clips[cat].keys())
        unused = sorted(all_keys - used_files[cat])
        if unused:
            print(f'  {folder}: ' + ', '.join(clips[cat][k] for k in unused))

    with open(OUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    print(f'\nWrote mapping -> {OUT_JSON}')


if __name__ == '__main__':
    main()
